from datetime import datetime
import os
from typing import Dict, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from auth import get_current_active_user
from db import get_db
from models import User, UserRole, ScheduleEntry, CollectionPeriod

router = APIRouter(prefix="/export", tags=["export"])


def standardize_time(time_str: str) -> str:
    if not time_str or ":" not in time_str:
        return time_str
    try:
        hours, minutes = map(int, time_str.split(":"))
        return f"{hours:02d}:{minutes:02d}"
    except ValueError:
        return time_str


def _build_schedule_string(entry: ScheduleEntry) -> str:
    if entry.status == "shift":
        meta = entry.meta or {}
        start = meta.get("shiftStart", "")
        end = meta.get("shiftEnd", "")
        if start and end:
            return f"{start}-{end}"
        return ""
    elif entry.status == "split":
        meta = entry.meta or {}
        s1 = meta.get("splitStart1", "")
        e1 = meta.get("splitEnd1", "")
        s2 = meta.get("splitStart2", "")
        e2 = meta.get("splitEnd2", "")
        if s1 and e1 and s2 and e2:
            return f"{s1}-{e1} {s2}-{e2}"
        return ""
    elif entry.status == "dayoff":
        return "выходной"
    elif entry.status == "vacation":
        return ""
    return ""


def _generate_excel_file(data: Dict) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "ИГ"

    all_dates = set()
    for user_info in data["data"].values():
        all_dates.update(user_info.get("schedule", {}).keys())
        all_dates.update(user_info.get("vacation_work", {}).keys())

    date_columns = sorted(all_dates, key=lambda x: datetime.strptime(x, "%Y-%m-%d"))

    headers = ["Группа", "ФИО", "Сумма часов"]
    for date_str in date_columns:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        headers.append(date_obj.strftime("%d.%b"))
        headers.append("")
    headers.extend(["", "Норма часов", "Доступность", "", "Доп. перерыв", "Комментарий"])
    ws.append(headers)

    N = len(date_columns)
    for i in range(N):
        start_col = 4 + i * 2
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1)

    blank_col = 3 + 2 * N + 1
    norm_start = blank_col + 1
    norm_end = norm_start
    avail_start_col = norm_end + 1
    avail_end_col = avail_start_col + 1
    ws.merge_cells(
        start_row=1,
        start_column=avail_start_col,
        end_row=1,
        end_column=avail_end_col
    )

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30

    sorted_users = sorted(
        data["data"].items(),
        key=lambda item: item[1]["full_name"].lower()
    )

    total_users = len(sorted_users)

    for idx, (user_id, user_data) in enumerate(sorted_users):
        group = user_data["alliance"]
        full_name = user_data["full_name"]
        vacation_work = user_data.get("vacation_work", {})
        schedule = user_data.get("schedule", {})

        avail_start, avail_end = "", ""

        shift_data = {}
        break_data = {}

        for date_str in date_columns:
            is_vacation = date_str in vacation_work
            value = schedule.get(date_str, "")

            if is_vacation:
                shift_start, shift_end = "отпуск", ""
                break_start, break_end = "", ""
            elif not value:
                shift_start, shift_end, break_start, break_end = "", "", "", ""
            elif "-" not in value:
                shift_start, shift_end, break_start, break_end = value, "", "", ""
            else:
                intervals = value.split()
                if len(intervals) == 1:
                    parts = intervals[0].split("-")
                    if len(parts) == 2:
                        shift_start, shift_end = parts
                        shift_start = standardize_time(shift_start.strip())
                        shift_end = standardize_time(shift_end.strip())
                        break_start, break_end = "", ""
                    else:
                        shift_start, shift_end, break_start, break_end = value, "", "", ""
                elif len(intervals) == 2:
                    parts1 = intervals[0].split("-")
                    parts2 = intervals[1].split("-")
                    if len(parts1) == 2 and len(parts2) == 2:
                        start1, end1 = parts1
                        start2, end2 = parts2
                        shift_start = standardize_time(start1.strip())
                        shift_end = standardize_time(end2.strip())
                        break_start = standardize_time(end1.strip())
                        break_end = standardize_time(start2.strip())
                    else:
                        shift_start, shift_end, break_start, break_end = value, "", "", ""
                else:
                    shift_start, shift_end, break_start, break_end = value, "", "", ""

            shift_data[date_str] = (shift_start, shift_end)
            break_data[date_str] = (break_start, break_end)

        row1 = [group, full_name, ""]
        for date_str in date_columns:
            row1.extend(shift_data[date_str])
        row1.extend(["", "", avail_start, avail_end, "", "", ""])
        ws.append(row1)

        row2 = ["Длительный перерыв", "", ""]
        for date_str in date_columns:
            row2.extend(break_data[date_str])

        total_columns = len(headers)
        current_len = len(row2)
        remaining = total_columns - current_len
        row2.extend([""] * remaining)
        ws.append(row2)

        break_row_idx = ws.max_row
        if idx < total_users - 1:
            ws.row_dimensions[break_row_idx].hidden = True
        ws.merge_cells(
            start_row=break_row_idx,
            start_column=1,
            end_row=break_row_idx,
            end_column=2
        )

        for row_idx in [ws.max_row - 1, ws.max_row]:
            for col_idx in range(4, 4 + 2 * N, 2):
                for offset in [0, 1]:
                    cell = ws.cell(row=row_idx, column=col_idx + offset)
                    if isinstance(cell.value, str) and ":" in cell.value:
                        try:
                            hours, minutes = map(int, cell.value.split(":"))
                            cell.value = f"{hours:02d}:{minutes:02d}"
                            cell.number_format = "h:mm"
                        except ValueError:
                            pass
            if row_idx == ws.max_row - 1:
                for col in [avail_start_col, avail_end_col]:
                    cell = ws.cell(row=row_idx, column=col)
                    if isinstance(cell.value, str) and ":" in cell.value:
                        try:
                            hours, minutes = map(int, cell.value.split(":"))
                            cell.value = f"{hours:02d}:{minutes:02d}"
                            cell.number_format = "h:mm"
                        except ValueError:
                            pass

    file_path = "temp_schedule.xlsx"
    wb.save(file_path)
    return file_path


@router.get("/schedule")
def export_schedule(
    period_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    if current_user.role not in (UserRole.ADMIN, UserRole.MANAGER):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    # Если period_id не указан, берем текущий открытый период
    if period_id:
        period = db.query(CollectionPeriod).filter(CollectionPeriod.id == period_id).first()
        if not period:
            raise HTTPException(status_code=404, detail="Период не найден")
        # Проверяем доступ к периоду
        if period.alliance != current_user.alliance:
            raise HTTPException(status_code=403, detail="Нет доступа к этому периоду")
    else:
        period = db.query(CollectionPeriod).filter(
            CollectionPeriod.is_open == True,
            CollectionPeriod.alliance == current_user.alliance
        ).first()
        if not period:
            raise HTTPException(status_code=400, detail="Нет активного периода сбора")

    query = db.query(User).filter(
        User.is_verified == True,
        User.alliance == current_user.alliance
    )

    users = query.all()

    export_data = {"data": {}}
    for user in users:
        entries = db.query(ScheduleEntry).filter(
            ScheduleEntry.user_id == user.id,
            ScheduleEntry.period_id == period.id
        ).all()

        schedule_dict = {}
        vacation_dict = {}
        for entry in entries:
            day_str = entry.day.isoformat()
            if entry.status == "vacation":
                vacation_dict[day_str] = True
            else:
                schedule_str = _build_schedule_string(entry)
                if schedule_str:
                    schedule_dict[day_str] = schedule_str

        export_data["data"][str(user.id)] = {
            "alliance": user.alliance or "",
            "full_name": user.full_name or user.email,
            "availability": "",  # пока пусто
            "vacation_work": vacation_dict,
            "schedule": schedule_dict
        }

    if not export_data["data"]:
        raise HTTPException(status_code=404, detail="Нет данных для экспорта")

    try:
        file_path = _generate_excel_file(export_data)
        return FileResponse(
            file_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f"schedule_{period.period_start}_{period.period_end}.xlsx"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации файла: {str(e)}")


@router.on_event("shutdown")
def cleanup():
    if os.path.exists("temp_schedule.xlsx"):
        os.remove("temp_schedule.xlsx")